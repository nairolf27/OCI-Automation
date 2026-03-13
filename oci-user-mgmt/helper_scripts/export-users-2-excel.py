"""
Filename: export_domains.py
Description: Exports users from a JSON domain file into a single Excel workbook,
             with one sheet per domain and columns: first_name, last_name, profile, groups.
             Only real user accounts (state=present, username containing @) are exported.
Author: Florian NOEL
Date: 13/03/2026
"""

import json
import logging
import os
import sys
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# ---------------------------------------------------------------------------
# Logging setup
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler(Path("logs") / "export_domains.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Environment variables
# ---------------------------------------------------------------------------

load_dotenv()

INPUT_FILE = os.getenv("INPUT_FILE", "domains.json")
OUTPUT_FILE = os.getenv("OUTPUT_FILE", "exports/domains_export.xlsx")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SHEET_FIELDNAMES = ["first_name", "last_name", "profile", "groups"]

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="2F5496")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
DATA_FONT = Font(name="Arial", size=11)
COLUMN_WIDTH = 25


# ---------------------------------------------------------------------------
# Functions
# ---------------------------------------------------------------------------


def load_json(file_path: str) -> dict:
    """
    Load and parse a JSON file from disk.

    :param file_path: Path to the JSON file to read.
    :return: Parsed content as a dictionary.
    :raises SystemExit: If the file is not found or contains invalid JSON.
    """
    path = Path(file_path)
    if not path.exists():
        logger.error("Input file not found: %s", file_path)
        sys.exit(1)

    with path.open(encoding="utf-8") as f:
        try:
            data = json.load(f)
            logger.info("Loaded JSON file: %s", file_path)
            return data
        except json.JSONDecodeError as exc:
            logger.error("Invalid JSON in %s: %s", file_path, exc)
            sys.exit(1)


def apply_header_style(sheet) -> None:
    """
    Apply professional formatting to the header row of a worksheet.

    :param sheet: The openpyxl worksheet to style.
    """
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT


def filter_exportable_users(users: list[dict], domain_name: str) -> list[dict]:
    """
    Retain only real, active user accounts from a domain's user list.

    Two conditions must both be met for a user to be exported:
    - state == "present": excludes deprovisioned or pending accounts.
    - "@" in username: excludes service accounts, which never carry an email-style username.

    :param users: Raw list of user dictionaries for a domain.
    :param domain_name: Domain name, used only for log context.
    :return: Filtered list of exportable user dictionaries.
    """
    exportable = [
        u for u in users
        if u.get("state") == "present" and "@" in u.get("username", "")
    ]

    skipped = len(users) - len(exportable)
    if skipped:
        logger.info(
            "Domain '%s': %d user(s) skipped (state != present or no @ in username)",
            domain_name,
            skipped,
        )

    return exportable


def write_domain_sheet(workbook: Workbook, domain_name: str, users: list[dict]) -> None:
    """
    Create and populate a worksheet for a single domain inside the workbook.

    Users are pre-filtered by filter_exportable_users before being written.
    Each row contains first_name, last_name, profile (may be empty), and groups
    (comma-separated). The sheet is named after the domain, truncated to Excel's
    31-character limit.

    :param workbook: The openpyxl Workbook to add the sheet to.
    :param domain_name: Name of the domain, used as the sheet title.
    :param users: List of user dictionaries for this domain.
    """
    sheet_title = domain_name[:31]
    sheet = workbook.create_sheet(title=sheet_title)

    sheet.append(SHEET_FIELDNAMES)
    apply_header_style(sheet)

    exportable_users = filter_exportable_users(users, domain_name)

    for user in exportable_users:
        groups = user.get("groups", [])
        row = [
            user.get("first_name", ""),
            user.get("last_name", ""),
            user.get("profile", ""),
            ", ".join(groups),
        ]
        sheet.append(row)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = DATA_FONT

    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = COLUMN_WIDTH

    logger.info(
        "Sheet created for domain '%s' with %d user(s)", domain_name, len(exportable_users)
    )


def export_all_domains(data: dict, output_path: Path) -> None:
    """
    Build a single Excel workbook with one sheet per domain, then save it to disk.

    The default sheet created by openpyxl is removed before saving so that only
    domain sheets are present in the final file.

    :param data: Full parsed JSON structure containing a 'domains' key.
    :param output_path: Destination path for the Excel file.
    """
    domains = data.get("domains", {})
    if not domains:
        logger.warning("No domains found in the input file.")
        return

    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)

    for domain_name, domain_data in domains.items():
        users = domain_data.get("users", [])
        logger.debug("Processing domain '%s' with %d user(s)", domain_name, len(users))
        write_domain_sheet(workbook, domain_name, users)

    workbook.save(output_path)
    logger.info("Workbook saved → %s (%d sheet(s))", output_path, len(workbook.sheetnames))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    """
    Entry point: loads the JSON input file and exports all domains into a single Excel workbook.
    """
    Path("logs").mkdir(exist_ok=True)
    logger.info(
        "Starting domain export — input: %s, output: %s", INPUT_FILE, OUTPUT_FILE
    )

    data = load_json(INPUT_FILE)
    export_all_domains(data, Path(OUTPUT_FILE))

    logger.info("Export complete.")


if __name__ == "__main__":
    main()
