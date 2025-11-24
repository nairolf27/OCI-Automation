"""
OCI Security Lists and Network Security Groups Excel Exporter

This script exports Security Lists (SL) and Network Security Groups (NSG) 
from Oracle Cloud Infrastructure (OCI) to an Excel file.

Requirements:
- oci (Oracle Cloud Infrastructure SDK)
- openpyxl (for Excel file manipulation)
- pandas (for data manipulation)

Install with: pip install oci openpyxl pandas
"""

import oci
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import sys

# ============================================================================
# CONFIGURABLE FUNCTION FOR IP INFO
# ============================================================================
def get_ip_info(ip_address):
    """
    Custom function to retrieve information about an IP address or CIDR.
    
    This function can be customized to:
    - Query an internal IPAM system
    - Look up DNS records
    - Check against known IP ranges
    - Add any custom logic you need
    
    Args:
        ip_address: IP address or CIDR block (e.g., "10.0.0.0/24", "192.168.1.1")
        
    Returns:
        String with information about the IP, or empty string
    """
    # Example implementation - customize this based on your needs
    
    # Example 1: Check for known ranges
    known_ranges = {
        "10.0.0.0/16": "Internal Network - Prod",
        "10.1.0.0/16": "Internal Network - Dev",
        "172.16.0.0/12": "Private Network Range",
        "192.168.0.0/16": "Management Network"
    }
    
    # Check if IP matches any known range
    for range_cidr, description in known_ranges.items():
        if ip_address == range_cidr:
            return description
    
    # Example 2: Check for special cases
    if ip_address == "0.0.0.0/0":
        return "Internet (All IPs)"
    
    # Example 3: You can add API calls here to query external systems
    # Example: query IPAM, DNS, or other inventory systems
    
    # Return empty string if no match
    return ""

# ============================================================================

# Initialize OCI configuration
# This will use the default config file at ~/.oci/config
config = oci.config.from_file("~/.oci/config", "cloudsh")

# Initialize OCI clients
virtual_network_client = oci.core.VirtualNetworkClient(config)
identity_client = oci.identity.IdentityClient(config)
integration_client = oci.integration.IntegrationInstanceClient(config)


def get_all_compartments(tenancy_id):
    """
    Retrieve all compartments in the tenancy (including nested compartments)
    
    Args:
        tenancy_id: OCI tenancy OCID
        
    Returns:
        List of all compartment objects
    """
    compartments = []
    try:
        # Add root compartment (tenancy)
        root = identity_client.get_compartment(tenancy_id).data
        compartments.append(root)
        
        # Get all compartments recursively
        def get_compartments_recursive(parent_id):
            try:
                comps = identity_client.list_compartments(
                    compartment_id=parent_id,
                    compartment_id_in_subtree=True
                ).data
                for comp in comps:
                    if comp.lifecycle_state == 'ACTIVE':
                        compartments.append(comp)
            except Exception as e:
                print(f"Warning: Error listing compartments under {parent_id}: {e}")
        
        get_compartments_recursive(tenancy_id)
        return compartments
    except Exception as e:
        print(f"Error retrieving compartments: {e}")
        return [root] if root else []


def get_vcn_by_name(tenancy_id, vcn_name):
    """
    Retrieve VCN object by name (searches all compartments in tenancy)
    
    Args:
        tenancy_id: OCI tenancy OCID
        vcn_name: Name of the VCN to search for
        
    Returns:
        Tuple of (VCN object, compartment_id) or (None, None) if not found
    """
    try:
        print(f"  Searching for VCN '{vcn_name}' across all compartments...")
        compartments = get_all_compartments(tenancy_id)
        print(f"  Searching in {len(compartments)} compartment(s)...")
        
        for compartment in compartments:
            try:
                vcns = virtual_network_client.list_vcns(
                    compartment_id=compartment.id
                ).data
                for vcn in vcns:
                    if vcn.display_name == vcn_name:
                        print(f"  Found in compartment: {compartment.name}")
                        return vcn, compartment.id
            except Exception as e:
                # Skip compartments where we don't have permission
                continue
        
        print(f"  VCN '{vcn_name}' not found in any compartment")
        return None, None
    except Exception as e:
        print(f"Error searching for VCN {vcn_name}: {e}")
        return None, None


def get_security_lists(compartment_id, vcn_id):
    """
    Retrieve all Security Lists for a given VCN
    
    Args:
        compartment_id: OCI compartment OCID
        vcn_id: VCN OCID
        
    Returns:
        List of Security List objects
    """
    try:
        security_lists = virtual_network_client.list_security_lists(
            compartment_id=compartment_id,
            vcn_id=vcn_id
        ).data
        return security_lists
    except Exception as e:
        print(f"Error retrieving Security Lists: {e}")
        return []


def get_network_security_groups(compartment_id, vcn_id):
    """
    Retrieve all Network Security Groups for a given VCN
    
    Args:
        compartment_id: OCI compartment OCID
        vcn_id: VCN OCID
        
    Returns:
        List of NSG objects
    """
    try:
        nsgs = virtual_network_client.list_network_security_groups(
            compartment_id=compartment_id,
            vcn_id=vcn_id
        ).data
        return nsgs
    except Exception as e:
        print(f"Error retrieving NSGs: {e}")
        return []


def get_nsg_rules(nsg_id):
    """
    Retrieve all rules for a given Network Security Group
    
    Args:
        nsg_id: NSG OCID
        
    Returns:
        List of NSG security rules
    """
    try:
        rules = virtual_network_client.list_network_security_group_security_rules(
            network_security_group_id=nsg_id
        ).data
        return rules
    except Exception as e:
        print(f"Error retrieving NSG rules: {e}")
        return []


def get_oic_instances(compartment_id):
    """
    Retrieve all OIC instances in a compartment
    
    Args:
        compartment_id: OCI compartment OCID
        
    Returns:
        List of OIC instance objects
    """
    try:
        instances = integration_client.list_integration_instances(
            compartment_id=compartment_id
        ).data
        return instances
    except Exception as e:
        print(f"Error retrieving OIC instances: {e}")
        return []


def get_all_oic_instances(tenancy_id):
    """
    Retrieve all OIC instances across all compartments in the tenancy
    
    Args:
        tenancy_id: OCI tenancy OCID
        
    Returns:
        List of tuples (OIC instance, compartment_id)
    """
    all_instances = []
    try:
        compartments = get_all_compartments(tenancy_id)
        print(f"  Searching for OIC instances in {len(compartments)} compartment(s)...")
        
        for compartment in compartments:
            try:
                instances = integration_client.list_integration_instances(
                    compartment_id=compartment.id
                ).data
                for instance in instances:
                    if instance.lifecycle_state == 'ACTIVE':
                        all_instances.append((instance, compartment.id))
            except Exception as e:
                # Skip compartments where we don't have permission or OIC service not available
                continue
        
        return all_instances
    except Exception as e:
        print(f"Error searching for OIC instances: {e}")
        return []


def format_oic_network_access_rules(oic_instance):
    """
    Format OIC Network Access rules into structured data
    
    Args:
        oic_instance: OIC instance object
        
    Returns:
        List of dictionaries containing formatted network access rules
    """
    formatted_rules = []
    
    try:
        # Get network endpoint details
        if hasattr(oic_instance, 'network_endpoint_details') and oic_instance.network_endpoint_details:
            network_details = oic_instance.network_endpoint_details
            
            # Check if there are allowlisted IPs or VCNs
            if hasattr(network_details, 'allowlisted_http_ips') and network_details.allowlisted_http_ips:
                for ip in network_details.allowlisted_http_ips:
                    rule_dict = {
                        'OIC Instance': oic_instance.display_name,
                        'Type': 'Ingress',
                        'IP': ip,
                        'Info IP': get_ip_info(ip),
                        'Protocol': 'HTTPS',
                        'Source Port': 'All',
                        'Destination Port Range': '443',
                        'Description': 'Allowlisted HTTP/HTTPS IP',
                        'Comments': ''
                    }
                    formatted_rules.append(rule_dict)
            
            if hasattr(network_details, 'allowlisted_http_vcns') and network_details.allowlisted_http_vcns:
                for vcn_access in network_details.allowlisted_http_vcns:
                    vcn_id = vcn_access.id if hasattr(vcn_access, 'id') else 'Unknown'
                    allowlisted_ips = vcn_access.allowlisted_ips if hasattr(vcn_access, 'allowlisted_ips') else []
                    
                    if allowlisted_ips:
                        for ip in allowlisted_ips:
                            rule_dict = {
                                'OIC Instance': oic_instance.display_name,
                                'Type': 'Ingress',
                                'IP': ip,
                                'Info IP': get_ip_info(ip),
                                'Protocol': 'HTTPS',
                                'Source Port': 'All',
                                'Destination Port Range': '443',
                                'Description': f'Allowlisted VCN IP (VCN: {vcn_id})',
                                'Comments': ''
                            }
                            formatted_rules.append(rule_dict)
                    else:
                        rule_dict = {
                            'OIC Instance': oic_instance.display_name,
                            'Type': 'Ingress',
                            'IP': f'VCN: {vcn_id}',
                            'Info IP': 'Entire VCN allowed',
                            'Protocol': 'HTTPS',
                            'Source Port': 'All',
                            'Destination Port Range': '443',
                            'Description': 'Allowlisted VCN access',
                            'Comments': ''
                        }
                        formatted_rules.append(rule_dict)
            
            # If no specific rules and it's public
            if not formatted_rules and hasattr(network_details, 'network_endpoint_type'):
                if network_details.network_endpoint_type == 'PUBLIC':
                    rule_dict = {
                        'OIC Instance': oic_instance.display_name,
                        'Type': 'Ingress',
                        'IP': '0.0.0.0/0',
                        'Info IP': get_ip_info('0.0.0.0/0'),
                        'Protocol': 'HTTPS',
                        'Source Port': 'All',
                        'Destination Port Range': '443',
                        'Description': 'Public endpoint - No IP restrictions',
                        'Comments': ''
                    }
                    formatted_rules.append(rule_dict)
        
        # If still no rules found, add a default entry
        if not formatted_rules:
            rule_dict = {
                'OIC Instance': oic_instance.display_name,
                'Type': 'N/A',
                'IP': 'N/A',
                'Info IP': '',
                'Protocol': 'N/A',
                'Source Port': 'N/A',
                'Destination Port Range': 'N/A',
                'Description': 'No network access rules configured',
                'Comments': ''
            }
            formatted_rules.append(rule_dict)
            
    except Exception as e:
        print(f"    Warning: Error processing network access rules: {e}")
        rule_dict = {
            'OIC Instance': oic_instance.display_name,
            'Type': 'Error',
            'IP': 'N/A',
            'Info IP': '',
            'Protocol': 'N/A',
            'Source Port': 'N/A',
            'Destination Port Range': 'N/A',
            'Description': f'Error retrieving rules: {str(e)}',
            'Comments': ''
        }
        formatted_rules.append(rule_dict)
    
    return formatted_rules


def format_port_range(port_range):
    """
    Format port range into a string
    
    Args:
        port_range: Port range object with min and max attributes
        
    Returns:
        String representation of port range (e.g., "80", "80-443", "All")
    """
    if not port_range:
        return "All"
    if port_range.min == port_range.max:
        return str(port_range.min)
    return f"{port_range.min}-{port_range.max}"


def format_security_list_rules(security_list):
    """
    Format Security List rules into structured data
    
    Args:
        security_list: Security List object
        
    Returns:
        List of dictionaries containing formatted rule data
    """
    formatted_rules = []
    
    # Process ingress rules
    for rule in security_list.ingress_security_rules:
        source_port = "All"
        dest_port = "All"
        
        # Extract port information based on protocol
        if hasattr(rule, 'tcp_options') and rule.tcp_options:
            if rule.tcp_options.source_port_range:
                source_port = format_port_range(rule.tcp_options.source_port_range)
            if rule.tcp_options.destination_port_range:
                dest_port = format_port_range(rule.tcp_options.destination_port_range)
                
        if hasattr(rule, 'udp_options') and rule.udp_options:
            if rule.udp_options.source_port_range:
                source_port = format_port_range(rule.udp_options.source_port_range)
            if rule.udp_options.destination_port_range:
                dest_port = format_port_range(rule.udp_options.destination_port_range)
        
        ip_address = rule.source
        
        rule_dict = {
            'Name': security_list.display_name,
            'Type': 'Ingress',
            'IP': ip_address,
            'Info IP': get_ip_info(ip_address),
            'Protocol': get_protocol_name(rule.protocol),
            'Source Port': source_port,
            'Destination Port Range': dest_port,
            'Description': rule.description or '',
            'Comments': ''
        }
        
        formatted_rules.append(rule_dict)
    
    # Process egress rules
    for rule in security_list.egress_security_rules:
        source_port = "All"
        dest_port = "All"
        
        # Extract port information based on protocol
        if hasattr(rule, 'tcp_options') and rule.tcp_options:
            if rule.tcp_options.source_port_range:
                source_port = format_port_range(rule.tcp_options.source_port_range)
            if rule.tcp_options.destination_port_range:
                dest_port = format_port_range(rule.tcp_options.destination_port_range)
                
        if hasattr(rule, 'udp_options') and rule.udp_options:
            if rule.udp_options.source_port_range:
                source_port = format_port_range(rule.udp_options.source_port_range)
            if rule.udp_options.destination_port_range:
                dest_port = format_port_range(rule.udp_options.destination_port_range)
        
        ip_address = rule.destination
        
        rule_dict = {
            'Name': security_list.display_name,
            'Type': 'Egress',
            'IP': ip_address,
            'Info IP': get_ip_info(ip_address),
            'Protocol': get_protocol_name(rule.protocol),
            'Source Port': source_port,
            'Destination Port Range': dest_port,
            'Description': rule.description or '',
            'Comments': ''
        }
        
        formatted_rules.append(rule_dict)
    
    return formatted_rules


def format_nsg_rules(nsg, rules):
    """
    Format NSG rules into structured data
    
    Args:
        nsg: NSG object
        rules: List of NSG security rules
        
    Returns:
        List of dictionaries containing formatted rule data
    """
    formatted_rules = []
    
    for rule in rules:
        source_port = "All"
        dest_port = "All"
        
        # Extract port information based on protocol
        if hasattr(rule, 'tcp_options') and rule.tcp_options:
            if rule.tcp_options.source_port_range:
                source_port = format_port_range(rule.tcp_options.source_port_range)
            if rule.tcp_options.destination_port_range:
                dest_port = format_port_range(rule.tcp_options.destination_port_range)
                
        if hasattr(rule, 'udp_options') and rule.udp_options:
            if rule.udp_options.source_port_range:
                source_port = format_port_range(rule.udp_options.source_port_range)
            if rule.udp_options.destination_port_range:
                dest_port = format_port_range(rule.udp_options.destination_port_range)
        
        # Determine IP based on direction
        ip_address = rule.source if rule.direction == 'INGRESS' else rule.destination
        
        rule_dict = {
            'Name': nsg.display_name,
            'Type': rule.direction.capitalize(),
            'IP': ip_address,
            'Info IP': get_ip_info(ip_address),
            'Protocol': get_protocol_name(rule.protocol),
            'Source Port': source_port,
            'Destination Port Range': dest_port,
            'Description': rule.description or '',
            'Comments': ''
        }
        
        formatted_rules.append(rule_dict)
    
    return formatted_rules


def get_protocol_name(protocol_number):
    """
    Convert protocol number to name
    
    Args:
        protocol_number: IP protocol number as string
        
    Returns:
        Protocol name (e.g., 'TCP', 'UDP', 'ICMP', or 'All')
    """
    protocol_map = {
        '6': 'TCP',
        '17': 'UDP',
        '1': 'ICMP',
        'all': 'All'
    }
    return protocol_map.get(protocol_number, f'Protocol {protocol_number}')


def style_worksheet(worksheet, title):
    """
    Apply styling to Excel worksheet
    
    Args:
        worksheet: openpyxl worksheet object
        title: Title for the worksheet
    """
    # Header styling
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    # Egress row styling (very light gray)
    egress_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    
    # Apply to first row (headers)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Apply gray background to Egress rows
    # Find the column index for 'Type'
    type_col_idx = None
    for idx, cell in enumerate(worksheet[1], 1):
        if cell.value == 'Type':
            type_col_idx = idx
            break
    
    if type_col_idx:
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            if row[type_col_idx - 1].value == 'Egress':
                for cell in row:
                    cell.fill = egress_fill
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def export_vcn_security_to_excel(tenancy_id, vcn_names, output_file):
    """
    Export Security Lists and NSGs for specified VCNs to Excel
    (Searches across all compartments in the tenancy)
    
    Args:
        tenancy_id: OCI tenancy OCID
        vcn_names: List of VCN names to export
        output_file: Path to output Excel file
    """
    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        
        for vcn_name in vcn_names:
            print(f"\nProcessing VCN: {vcn_name}")
            
            # Get VCN (searches all compartments)
            vcn, compartment_id = get_vcn_by_name(tenancy_id, vcn_name)
            if not vcn:
                print(f"  Warning: VCN '{vcn_name}' not found. Skipping...")
                continue
            
            print(f"  Found VCN: {vcn.display_name} (ID: {vcn.id})")
            
            # Process Security Lists
            security_lists = get_security_lists(compartment_id, vcn.id)
            print(f"  Found {len(security_lists)} Security List(s)")
            
            if security_lists:
                all_sl_data = []
                
                for sl in security_lists:
                    print(f"    Processing Security List: {sl.display_name}")
                    sl_rules = format_security_list_rules(sl)
                    all_sl_data.extend(sl_rules)
                
                if all_sl_data:
                    # Create DataFrame and write to Excel
                    df_sl = pd.DataFrame(all_sl_data)
                    # Reorder columns to match requested order
                    column_order = ['Name', 'Type', 'IP', 'Info IP', 'Protocol', 'Source Port', 'Destination Port Range', 'Description', 'Comments']
                    df_sl = df_sl[column_order]
                    sheet_name = f"{vcn_name[:25]}_SL"  # Limit sheet name to 31 chars
                    df_sl.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Style the worksheet
                    worksheet = writer.sheets[sheet_name]
                    style_worksheet(worksheet, f"{vcn_name} - Security Lists")
                    print(f"  Exported {len(all_sl_data)} Security List rules")
            
            # Process Network Security Groups
            nsgs = get_network_security_groups(compartment_id, vcn.id)
            print(f"  Found {len(nsgs)} Network Security Group(s)")
            
            if nsgs:
                all_nsg_data = []
                
                for nsg in nsgs:
                    print(f"    Processing NSG: {nsg.display_name}")
                    rules = get_nsg_rules(nsg.id)
                    nsg_data = format_nsg_rules(nsg, rules)
                    all_nsg_data.extend(nsg_data)
                
                if all_nsg_data:
                    # Create DataFrame and write to Excel
                    df_nsg = pd.DataFrame(all_nsg_data)
                    # Reorder columns to match requested order
                    column_order = ['Name', 'Type', 'IP', 'Info IP', 'Protocol', 'Source Port', 'Destination Port Range', 'Description', 'Comments']
                    df_nsg = df_nsg[column_order]
                    sheet_name = f"{vcn_name[:24]}_NSG"  # Limit sheet name to 31 chars
                    df_nsg.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Style the worksheet
                    worksheet = writer.sheets[sheet_name]
                    style_worksheet(worksheet, f"{vcn_name} - Network Security Groups")
                    print(f"  Exported {len(all_nsg_data)} NSG rules")
    
    print(f"\n✓ Export completed successfully: {output_file}")


def main():
    """
    Main function to execute the script
    """
    print("=" * 70)
    print("OCI Security Lists and NSG Exporter")
    print("=" * 70)
    
    # Get tenancy ID from config
    tenancy_id = config.get('tenancy')
    print(f"\nUsing tenancy ID: {tenancy_id}")
    print("Will search for VCNs across ALL compartments in the tenancy")
    
    # Get VCN names from user input
    print("\nEnter VCN names to analyze (comma-separated):")
    vcn_input = input("> ")
    vcn_names = [name.strip() for name in vcn_input.split(",")]
    
    # Get output filename
    print("\nEnter output Excel filename (default: oci_security_export.xlsx):")
    output_file = input("> ").strip()
    if not output_file:
        output_file = "oci_security_export.xlsx"
    
    if not output_file.endswith('.xlsx'):
        output_file += '.xlsx'
    
    # Execute export
    print(f"\nStarting export for {len(vcn_names)} VCN(s)...")
    export_vcn_security_to_excel(tenancy_id, vcn_names, output_file)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nExport cancelled by user.")
        sys.exit(0)
    except Exception as e:
        print(f"\n✗ Error occurred: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
